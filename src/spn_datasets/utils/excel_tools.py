# coding=UTF-8
"""This module provides a utility class for interacting with Excel files."""

from pathlib import Path
import openpyxl
from openpyxl import Workbook
from typing import List, Any


class ExcelTool:
    """A tool for creating and manipulating Excel (.xlsx) files."""

    def __init__(self, path: Path, book_name: str, sheet_name: str) -> None:
        """Initializes the ExcelTool.

        Args:
            path: The directory where the Excel file will be saved.
            book_name: The name of the Excel file (with or without .xlsx extension).
            sheet_name: The name of the sheet to work with.
        """
        super(ExcelTool, self).__init__()
        self.path = path
        book_path = Path(book_name)
        if book_path.suffix != ".xlsx":
            self.book_name = book_path.with_suffix(".xlsx")
        else:
            self.book_name = book_path
        self.excel_loc = self.path / self.book_name
        self.sheet_name = sheet_name

    def write_xls(self, value: List[List[Any]]) -> None:
        """Writes data to a new Excel file, overwriting it if it exists.

        Args:
            value: A 2D list of data to write to the sheet.
        """
        if not self.path.exists():
            self.path.mkdir(parents=True, exist_ok=True)

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = self.sheet_name

        for row_idx, row_data in enumerate(value, 1):
            for col_idx, cell_value in enumerate(row_data, 1):
                sheet.cell(row=row_idx, column=col_idx, value=cell_value)

        workbook.save(self.excel_loc)
        print("xlsx file created successfully!")

    def append_to_xls(self, value: List[List[Any]]) -> None:
        """Appends data to an existing Excel file, or creates it if it doesn't exist.

        Args:
            value: A 2D list of data to append to the sheet.
        """
        if not self.excel_loc.exists():
            self.write_xls(value)
            return

        workbook = openpyxl.load_workbook(self.excel_loc)
        sheet = workbook.active

        last_row = sheet.max_row
        for row_idx, row_data in enumerate(value, 1):
            for col_idx, cell_value in enumerate(row_data, 1):
                sheet.cell(row=last_row + row_idx, column=col_idx, value=cell_value)

        workbook.save(self.excel_loc)
        print("Data appended to xlsx file successfully!")

    def read_excel_xls(self) -> None:
        """Reads and prints the content of the Excel file."""
        if not self.excel_loc.exists():
            print(f"Error: File '{self.excel_loc}' not found.")
            return

        workbook = openpyxl.load_workbook(self.excel_loc)
        sheet = workbook.active

        for row in sheet.iter_rows():
            print("\t".join(str(cell.value) if cell.value is not None else "" for cell in row))

    def append_blank_lines(self, num_lines: int) -> None:
        """Appends a specified number of blank lines to the Excel file.

        Args:
            num_lines: The number of blank lines to add.
        """
        if not self.excel_loc.exists():
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = self.sheet_name
        else:
            workbook = openpyxl.load_workbook(self.excel_loc)
            sheet = workbook.active

        last_row = sheet.max_row
        for i in range(num_lines):
            sheet.cell(row=last_row + i + 1, column=1, value=" ")
            sheet.cell(row=last_row + i + 1, column=2, value=" ")

        workbook.save(self.excel_loc)
        print(f"{num_lines} blank lines appended successfully!")

    def append_value_and_blank_lines(self, value: List[List[Any]], num_blank_lines: int) -> None:
        """Appends data to the Excel file, followed by a specified number of blank lines.

        Args:
            value: The 2D list of data to append.
            num_blank_lines: The number of blank lines to add after the data.
        """
        self.append_to_xls(value)
        self.append_blank_lines(num_blank_lines)
