# coding=UTF-8
from pathlib import Path
import openpyxl
from openpyxl import Workbook


class ExcelTool:
    def __init__(self, path, book_name_xls, sheet_name_xls):
        super(ExcelTool, self).__init__()
        self.path = Path(path)
        # Ensure the file extension is .xlsx
        book_path = Path(book_name_xls)
        if book_path.suffix != '.xlsx':
            self.book_name_xls = book_path.with_suffix('.xlsx')
        else:
            self.book_name_xls = book_path
        self.excel_loc = self.path / self.book_name_xls
        self.sheet_name_xls = sheet_name_xls

    def write_xls(self, value):
        if not self.path.exists():
            self.path.mkdir(parents=True, exist_ok=True)

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = self.sheet_name_xls

        for row_idx, row_data in enumerate(value, 1):
            for col_idx, cell_value in enumerate(row_data, 1):
                sheet.cell(row=row_idx, column=col_idx, value=cell_value)

        workbook.save(self.excel_loc)
        print("xlsx file created successfully!")

    def write_xls_append(self, value):
        if not self.excel_loc.exists():
            # If the file doesn't exist, create it with the initial data
            self.write_xls(value)
            return

        workbook = openpyxl.load_workbook(self.excel_loc)
        sheet = workbook.active

        last_row = sheet.max_row
        for row_idx, row_data in enumerate(value, 1):
            for col_idx, cell_value in enumerate(row_data, 1):
                sheet.cell(row=last_row + row_idx, column=col_idx, value=cell_value)

        workbook.save(self.excel_loc)
        print("Data appended to xlsx file successfully!")

    def read_excel_xls(self):
        if not self.excel_loc.exists():
            print(f"Error: File '{self.excel_loc}' not found.")
            return

        workbook = openpyxl.load_workbook(self.excel_loc)
        sheet = workbook.active

        for row in sheet.iter_rows():
            print("\t".join(str(cell.value) if cell.value is not None else "" for cell in row))

    def write_append_blankline(self, length):
        """

        :param length:
            -------
            length: the size of add blankline
        :return:
            ------
            none
        """
        if not self.excel_loc.exists():
            # If the file doesn't exist, create it with blank lines
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = self.sheet_name_xls
        else:
            workbook = openpyxl.load_workbook(self.excel_loc)
            sheet = workbook.active

        last_row = sheet.max_row
        for i in range(length):
            sheet.cell(row=last_row + i + 1, column=1, value=" ")
            sheet.cell(row=last_row + i + 1, column=2, value=" ")

        workbook.save(self.excel_loc)
        print("Blank lines appended successfully!")

    def write_append_value_blank(self, value, length):
        self.write_xls_append(value)
        self.write_append_blankline(length)
